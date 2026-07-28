"""
Tracks all job applications in an Excel spreadsheet.
File: ~/job-automation/Job_Applications_Tracker.xlsx
"""

import logging
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

TRACKER_PATH = os.path.join(os.path.dirname(__file__), "Job_Applications_Tracker.xlsx")

HEADERS = [
    "Date Applied", "Job Title", "Company", "Location",
    "Platform", "Status", "Job URL", "Resume Used", "Notes"
]

STATUS_COLORS = {
    "applied":                  "C8E6C9",  # green
    "skipped:no_easy_apply":    "FFF9C4",  # yellow
    "skipped:manual_apply_required": "FFF9C4",
    "failed":                   "FFCDD2",  # red
    "dry_run":                  "E3F2FD",  # blue
}

PLATFORM_COLORS = {
    "linkedin":   "BBDEFB",
    "indeed":     "C5CAE9",
    "jobs.ie":    "FFCCBC",
    "gradireland":"DCEDC8",
}


def _header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="1A237E")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _status_label(status: str) -> str:
    labels = {
        "applied": "Applied",
        "skipped:no_easy_apply": "No Easy Apply",
        "skipped:manual_apply_required": "Manual Apply",
        "skipped:no_selenium": "Selenium Missing",
        "dry_run": "Dry Run",
        "unknown": "Unknown",
    }
    if status in labels:
        return labels[status]
    if status.startswith("failed"):
        return "Failed"
    return status.replace("_", " ").title()


def _init_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Header row
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _header_style(cell)

    # Column widths
    widths = [16, 32, 24, 16, 12, 18, 50, 36, 24]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    return wb


def update_tracker(jobs: list[dict], statuses: dict[str, str], tailored_resumes: dict[str, str]):
    """
    Add applied jobs to the Excel tracker.
    Creates the file if it doesn't exist.
    """
    if os.path.exists(TRACKER_PATH):
        try:
            wb = load_workbook(TRACKER_PATH)
            ws = wb.active
        except Exception:
            wb = _init_workbook()
            ws = wb.active
    else:
        wb = _init_workbook()
        ws = wb.active

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    added = 0
    for job in jobs:
        job_id = job["id"]
        status = statuses.get(job_id, "unknown")
        resume_path = tailored_resumes.get(job_id, "Original Resume")
        resume_name = os.path.basename(resume_path) if resume_path else "Original Resume"
        status_label = _status_label(status)

        # Determine row color
        fill_color = None
        for key, color in STATUS_COLORS.items():
            if status.startswith(key):
                fill_color = color
                break
        if not fill_color:
            fill_color = "F5F5F5"

        row_data = [
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            job.get("platform", "").title(),
            status_label,
            job.get("url", ""),
            resume_name,
            "",  # Notes — filled manually
        ]

        next_row = ws.max_row + 1
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col, value=value)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 7))
            if col == 7:  # URL column — make it a hyperlink
                cell.hyperlink = value
                cell.font = Font(color="0000EE", underline="single")
            else:
                cell.font = Font(size=10)

        ws.row_dimensions[next_row].height = 18
        added += 1

    try:
        wb.save(TRACKER_PATH)
        logger.info(f"Tracker updated: {added} rows added → {TRACKER_PATH}")
    except Exception as e:
        logger.error(f"Could not save tracker: {e}")

    return TRACKER_PATH
